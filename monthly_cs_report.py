"""Monthly CS report -- AIA-paid customers grouped by their 1st-payment month.

Standalone script (imports main.py so all activity-score / cadence / customer-
status math is the exact same code the live dashboard already uses -- one
source of truth, not a second copy to keep in sync). Meant to be run by an
n8n Schedule Trigger + SSH "Execute a command" node on/after the 5th of each
month: `python3 monthly_cs_report.py`.

Builds a 3-sheet workbook:
  - "{Mon YY} Paid" for current-2m (1st payment 2 calendar months ago)
  - "{Mon YY} Paid" for current-1m (1st payment 1 calendar month ago)
  - "Overall" -- 4 pivot tables (Stage x CSM, Activity Status x CSM, one pair
    per cohort month)
and prints exactly one line to stdout:
  REPORT_JSON={"file_path": "...", "filename": "...", "subject": "...", "overall_html": "..."}
so a downstream n8n Code node can extract it with /REPORT_JSON=(.+)/, the same
sentinel-line convention run_snapshot.sh already uses for LOCAL_PDF=.
Everything main.py prints during its own data load is redirected to stderr so
it never lands on this stdout line.
"""
import sys, os, io, json, contextlib
from datetime import date

_HERE = os.path.dirname(os.path.abspath(__file__))
os.chdir(_HERE)                 # so main.py's load_dotenv()/relative imports resolve
sys.path.insert(0, _HERE)

_buf = io.StringIO()
with contextlib.redirect_stdout(_buf):
    import main as m
for _line in _buf.getvalue().splitlines():
    print(f"[main] {_line}", file=sys.stderr)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HS_BASE = "https://app-na2.hubspot.com/contacts/39668252/record/0-3/"
STAGES   = ["Integration Done", "Product Blocked", "Renewal Done",
            "Ready for Renewal", "CS Parked", "Churned"]
STATUSES = ["Active", "Risk of Churn", "Inactive"]
YELLOW     = "FFFF00"
TOTAL_FILL = "D9E1F2"
HEADER_FILL = "1B2430"

ROW_COLS = ["Deal Name", "record_id", "Payment Date", "Integration Date", "Deal Stage",
            "GM", "CSM", "Active Days (28D)", "Activity Score (28D)", "Cadence",
            "Activity Stage", "Churned Reason", "CS Parked Reason",
            "Product Blocked Reason", "Comments"]


def month_cohort(aia, first_of_month):
    start = pd.Timestamp(first_of_month)
    end = start + pd.DateOffset(months=1) - pd.Timedelta(days=1)
    pay = pd.to_datetime(aia["payment_date"], errors="coerce")
    return aia[(pay >= start) & (pay <= end)].copy()


def build_rows(cohort, ev_lu, scores, today):
    rows = []
    for _, row in cohort.iterrows():
        email = row.get("login_email_id", "")
        acct = m._acct_for(email)
        active_days, _streak = m._usage_28(email, ev_lu)
        activity_score = int(scores.get(acct, 0)) if acct else 0
        intd = row.get("integration_done_date")
        cad = row.get("cadence") or "Monthly"
        if pd.notna(intd):
            intd_ts = pd.Timestamp(intd).normalize()
            dsince = (today - intd_ts).days
            status = m._customer_status_m(acct, intd_ts, cad, dsince, today) or ""
        else:
            status = ""
        pay = row.get("payment_date")
        rows.append({
            "Deal Name":               row.get("deal_name", "") or "",
            "record_id":               row.get("record_id", ""),
            "Payment Date":            pd.Timestamp(pay).strftime("%d-%b-%y") if pd.notna(pay) else "",
            "Integration Date":        pd.Timestamp(intd).strftime("%d-%b-%y") if pd.notna(intd) else "",
            "Deal Stage":              row.get("deal_stage", "") or "",
            "GM":                      row.get("deal_owner", "") or "",
            "CSM":                     row.get("cs_owner", "") or "",
            "Active Days (28D)":       active_days,
            "Activity Score (28D)":    activity_score,
            "Cadence":                 cad,
            "Activity Stage":          status,
            "Churned Reason":          row.get("churned_reason", "") or "",
            "CS Parked Reason":        row.get("cs_parked_reason", "") or "",
            "Product Blocked Reason":  row.get("blocked_reason", "") or "",
            "Comments":                "",
        })
    return pd.DataFrame(rows, columns=ROW_COLS)


def pivot(df_rows, label_col, cats, split_col):
    csms = sorted(df_rows["CSM"].dropna().unique()) if len(df_rows) else []
    tbl = pd.DataFrame({label_col: cats})
    for csm in csms:
        sub = df_rows[df_rows["CSM"] == csm]
        tbl[csm] = [int((sub[split_col] == c).sum()) for c in cats]
    tbl.loc[len(tbl)] = ["Total"] + [int(tbl[c].sum()) for c in csms]
    return tbl


def write_month_sheet(wb, title, df_rows):
    ws = wb.create_sheet(title[:31])
    headers = [c for c in ROW_COLS if c != "record_id"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    link_font = Font(color="2C5AA0", underline="single")
    name_col = headers.index("Deal Name") + 1
    for _, r in df_rows.iterrows():
        ws.append([r[h] for h in headers])
        rix = ws.max_row
        if r.get("record_id"):
            cell = ws.cell(row=rix, column=name_col)
            cell.hyperlink = HS_BASE + str(r["record_id"])
            cell.font = link_font
    widths = {"Deal Name": 42, "Payment Date": 14, "Integration Date": 16, "Deal Stage": 18,
              "GM": 20, "CSM": 20, "Active Days (28D)": 16, "Activity Score (28D)": 18,
              "Cadence": 12, "Activity Stage": 14, "Churned Reason": 26,
              "CS Parked Reason": 26, "Product Blocked Reason": 26, "Comments": 34}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 16)
    ws.freeze_panes = "A2"
    return ws


def write_pivot(ws, start_row, section_title, tbl, label_col):
    ws.cell(row=start_row, column=1, value=section_title).font = Font(bold=True, size=12)
    r = start_row + 1
    hcell = ws.cell(row=r, column=1, value=label_col)
    hcell.font = Font(bold=True)
    hcell.fill = PatternFill("solid", fgColor=YELLOW)
    for j, csm in enumerate(tbl.columns[1:], start=2):
        ws.cell(row=r, column=j, value=csm).font = Font(bold=True)
    r += 1
    for _, row in tbl.iterrows():
        is_total = row[label_col] == "Total"
        for j, col in enumerate(tbl.columns, start=1):
            c = ws.cell(row=r, column=j, value=row[col])
            if is_total:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            if j > 1:
                c.alignment = Alignment(horizontal="center")
        r += 1
    return r + 1


def write_overall_sheet(wb, m2_label, m2_stage, m2_status, m1_label, m1_stage, m1_status):
    ws = wb.create_sheet("Overall")
    r = 1
    r = write_pivot(ws, r, f"{m2_label} Paid — Stage", m2_stage, "Stage")
    r = write_pivot(ws, r, f"{m2_label} Paid — Activity Status", m2_status, "Activity Status")
    r = write_pivot(ws, r, f"{m1_label} Paid — Stage", m1_stage, "Stage")
    r = write_pivot(ws, r, f"{m1_label} Paid — Activity Status", m1_status, "Activity Status")
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18
    return ws


def _pivot_html(title, tbl, label_col):
    csms = list(tbl.columns[1:])
    th = "".join(
        f'<th style="padding:6px 10px;background:#1B2430;color:#ffffff;'
        f'font-size:12px;text-align:center">{c}</th>' for c in csms)
    body = ""
    for _, row in tbl.iterrows():
        is_total = row[label_col] == "Total"
        bg = "#D9E1F2" if is_total else "#ffffff"
        weight = "700" if is_total else "400"
        tds = "".join(
            f'<td style="padding:6px 10px;text-align:center;font-weight:{weight};'
            f'border:1px solid #e0e0e0">{row[c]}</td>' for c in csms)
        body += (f'<tr style="background:{bg}"><td style="padding:6px 10px;'
                 f'font-weight:{weight};border:1px solid #e0e0e0">{row[label_col]}'
                 f'</td>{tds}</tr>')
    return (f'<p style="font-weight:700;margin:18px 0 4px;font-size:14px;'
            f'font-family:Segoe UI,Arial,sans-serif">{title}</p>'
            f'<table style="border-collapse:collapse;font-family:Segoe UI,Arial,sans-serif;'
            f'font-size:13px;width:100%;max-width:640px">'
            f'<tr><th style="padding:6px 10px;background:{"#" + YELLOW};color:#000000;'
            f'text-align:left;border:1px solid #e0e0e0">{label_col}</th>{th}</tr>{body}</table>')


def build_overall_html(m2_label, m2_stage, m2_status, m1_label, m1_stage, m1_status):
    return (_pivot_html(f"{m2_label} Paid — Stage", m2_stage, "Stage")
            + _pivot_html(f"{m2_label} Paid — Activity Status", m2_status, "Activity Status")
            + _pivot_html(f"{m1_label} Paid — Stage", m1_stage, "Stage")
            + _pivot_html(f"{m1_label} Paid — Activity Status", m1_status, "Activity Status"))


def main_run():
    today = pd.Timestamp(date.today()).normalize()
    this_month = today.replace(day=1)
    m1_start = this_month - pd.DateOffset(months=1)   # current - 1 month
    m2_start = this_month - pd.DateOffset(months=2)   # current - 2 months
    m1_label = m1_start.strftime("%b %y")
    m2_label = m2_start.strftime("%b %y")

    aia = m._AIA
    ev_lu = m._recent_event_lookup()
    scores = m._activity_scores()

    m1_rows = build_rows(month_cohort(aia, m1_start), ev_lu, scores, today)
    m2_rows = build_rows(month_cohort(aia, m2_start), ev_lu, scores, today)

    m1_stage,  m1_status = pivot(m1_rows, "Stage", STAGES, "Deal Stage"), pivot(m1_rows, "Activity Status", STATUSES, "Activity Stage")
    m2_stage,  m2_status = pivot(m2_rows, "Stage", STAGES, "Deal Stage"), pivot(m2_rows, "Activity Status", STATUSES, "Activity Stage")

    wb = Workbook()
    wb.remove(wb.active)
    write_month_sheet(wb, f"{m2_label} Paid", m2_rows)
    write_month_sheet(wb, f"{m1_label} Paid", m1_rows)
    write_overall_sheet(wb, m2_label, m2_stage, m2_status, m1_label, m1_stage, m1_status)

    out_dir = os.path.join(_HERE, "exports")
    os.makedirs(out_dir, exist_ok=True)
    filename = f"CS_Monthly_Report_{today.strftime('%b_%Y')}.xlsx"
    file_path = os.path.join(out_dir, filename)
    wb.save(file_path)

    # In production this runs INSIDE the aia-dashboard container (`docker exec`),
    # where file_path is /app/exports/... . docker-compose bind-mounts ./exports
    # to /app/exports, so the same file is on the HOST at HOST_EXPORT_DIR --
    # which is the path n8n's SSH download node needs (it reads the host FS).
    host_dir = os.environ.get("HOST_EXPORT_DIR", "/opt/taipy-dashboard/exports")
    result = {
        "file_path": file_path,                                   # container-side
        "host_file_path": f"{host_dir.rstrip('/')}/{filename}",   # host-side (n8n)
        "filename": filename,
        "subject": f"CS Monthly Report — {today.strftime('%b %Y')}",
        "overall_html": build_overall_html(m2_label, m2_stage, m2_status, m1_label, m1_stage, m1_status),
    }
    print("REPORT_JSON=" + json.dumps(result))


if __name__ == "__main__":
    main_run()
